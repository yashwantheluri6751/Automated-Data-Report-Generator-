"""
Automated Data Report Generator
================================
Reads raw Excel sales data, computes KPIs automatically,
generates AI-style plain-English insights, and produces
a formatted, ready-to-share Excel report in under 30 seconds.

Author : Yashwanth Eluri
GitHub : https://github.com/yashwantheluri6751
"""

import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

# ── CONFIG ────────────────────────────────────────────────────────────────────

INPUT_FILE  = "sales_data.xlsx"       # Raw sales input
OUTPUT_FILE = "automated_report.xlsx" # Generated report

# ── AI INSIGHT ENGINE ─────────────────────────────────────────────────────────

class AIInsightEngine:
    """
    Rule-based AI engine that generates plain-English business insights.
    Mimics what an analyst (or LLM) would write after looking at the numbers.
    """

    def growth_insight(self, current, previous, metric_name):
        if previous == 0:
            return f"{metric_name} recorded for the first time this period."
        growth = ((current - previous) / abs(previous)) * 100
        direction = "grew" if growth > 0 else "declined"
        magnitude = "sharply" if abs(growth) > 20 else ("steadily" if abs(growth) > 5 else "slightly")
        return f"{metric_name} {direction} {magnitude} by {abs(growth):.1f}% compared to previous period."

    def top_performer_insight(self, df, group_col, value_col):
        top    = df.groupby(group_col)[value_col].sum().idxmax()
        top_val= df.groupby(group_col)[value_col].sum().max()
        total  = df[value_col].sum()
        share  = (top_val / total) * 100
        return f"{top} is the top performer, contributing {share:.1f}% of total {value_col.replace('_',' ')}."

    def underperformer_insight(self, df, group_col, value_col):
        bottom = df.groupby(group_col)[value_col].sum().idxmin()
        avg    = df.groupby(group_col)[value_col].sum().mean()
        actual = df.groupby(group_col)[value_col].sum().min()
        gap    = avg - actual
        return f"{bottom} is underperforming — {gap:,.0f} units below average. Consider promotional push."

    def trend_insight(self, monthly_series):
        if len(monthly_series) < 2:
            return "Insufficient data for trend analysis."
        recent   = monthly_series.iloc[-1]
        previous = monthly_series.iloc[-2]
        peak     = monthly_series.max()
        if recent == peak:
            return "Current month is the highest performing month on record — positive momentum."
        elif recent > previous:
            return f"Revenue recovering — up from last month. Monitor to confirm sustained growth."
        else:
            return f"Revenue dipped this month. Review pricing, inventory, or demand factors."

    def overall_health(self, total_revenue, total_units, avg_order_value):
        if avg_order_value > 5000:
            tier = "premium"
        elif avg_order_value > 2000:
            tier = "mid-market"
        else:
            tier = "volume-driven"
        return (
            f"Business processed {total_units:,} units generating "
            f"₹{total_revenue:,.0f} total revenue. "
            f"Average order value of ₹{avg_order_value:,.0f} indicates a {tier} sales profile."
        )

    def recommendation(self, df, region_col, revenue_col):
        by_region = df.groupby(region_col)[revenue_col].sum()
        top_region = by_region.idxmax()
        low_region = by_region.idxmin()
        return (
            f"Recommended actions: (1) Scale marketing in {top_region} — highest revenue region. "
            f"(2) Investigate {low_region} — lowest performing region may need targeted campaigns."
        )


# ── DATA LOADING ──────────────────────────────────────────────────────────────

def load_sales_data(filepath):
    """Load Excel data or generate sample data if file not found."""
    if os.path.exists(filepath):
        df = pd.read_excel(filepath)
        print(f"Loaded {len(df)} rows from {filepath}")
    else:
        print(f"'{filepath}' not found — generating sample sales data.")
        df = _generate_sample_data()
        df.to_excel(filepath, index=False)
        print(f"Sample data saved to '{filepath}' — edit it and re-run!")
    return df


def _generate_sample_data():
    """Realistic 120-row sample dataset across products, regions, months."""
    import random
    random.seed(42)

    products = ["Laptop", "Smartphone", "Tablet", "Headphones", "Smartwatch",
                "Monitor", "Keyboard", "Mouse", "Webcam", "Speaker"]
    regions  = ["North", "South", "East", "West"]
    months   = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

    rows = []
    for month_idx, month in enumerate(months):
        for product in products:
            units     = random.randint(10, 200)
            base_price= {"Laptop":55000,"Smartphone":25000,"Tablet":18000,
                         "Headphones":3500,"Smartwatch":8000,"Monitor":22000,
                         "Keyboard":2500,"Mouse":1200,"Webcam":4500,"Speaker":6000}
            price     = base_price[product] * random.uniform(0.9, 1.15)
            revenue   = units * price
            region    = random.choice(regions)
            rows.append({
                "month"      : month,
                "month_num"  : month_idx + 1,
                "product"    : product,
                "region"     : region,
                "units_sold" : units,
                "unit_price" : round(price, 2),
                "revenue"    : round(revenue, 2),
                "returns"    : random.randint(0, max(1, units // 10)),
                "salesperson": f"SP-{random.randint(1,8):02d}"
            })
    return pd.DataFrame(rows)


# ── KPI COMPUTATION ───────────────────────────────────────────────────────────

def compute_kpis(df):
    """Compute all KPIs needed for the report."""
    kpis = {}

    kpis["total_revenue"]     = df["revenue"].sum()
    kpis["total_units"]       = df["units_sold"].sum()
    kpis["total_returns"]     = df["returns"].sum()
    kpis["avg_order_value"]   = df["revenue"].sum() / len(df)
    kpis["return_rate"]       = (df["returns"].sum() / df["units_sold"].sum()) * 100

    # Monthly
    monthly = df.groupby(["month_num","month"]).agg(
        revenue=("revenue","sum"), units=("units_sold","sum")
    ).reset_index().sort_values("month_num")
    kpis["monthly"] = monthly

    # By product
    kpis["by_product"] = df.groupby("product").agg(
        revenue=("revenue","sum"), units=("units_sold","sum"), returns=("returns","sum")
    ).reset_index().sort_values("revenue", ascending=False)

    # By region
    kpis["by_region"] = df.groupby("region").agg(
        revenue=("revenue","sum"), units=("units_sold","sum")
    ).reset_index().sort_values("revenue", ascending=False)

    # Growth (last month vs second last)
    if len(monthly) >= 2:
        kpis["last_month_rev"]     = monthly.iloc[-1]["revenue"]
        kpis["prev_month_rev"]     = monthly.iloc[-2]["revenue"]
        kpis["last_month_units"]   = monthly.iloc[-1]["units"]
        kpis["prev_month_units"]   = monthly.iloc[-2]["units"]
    else:
        kpis["last_month_rev"]   = kpis["total_revenue"]
        kpis["prev_month_rev"]   = kpis["total_revenue"]
        kpis["last_month_units"] = kpis["total_units"]
        kpis["prev_month_units"] = kpis["total_units"]

    return kpis


# ── EXCEL REPORT BUILDER ──────────────────────────────────────────────────────

def build_report(df, kpis, ai, output_path):
    wb = Workbook()
    _sheet_executive_summary(wb, kpis, ai, df)
    _sheet_monthly_trend(wb, kpis)
    _sheet_product_analysis(wb, kpis)
    _sheet_raw_data(wb, df)
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    wb.save(output_path)


def _border():
    s = Side(style="thin", color="DDDDDD")
    return Border(left=s, right=s, top=s, bottom=s)


def _hdr(ws, row, bg="1A237E", fg="FFFFFF"):
    fill = PatternFill("solid", fgColor=bg)
    font = Font(bold=True, color=fg, size=11)
    for cell in ws[row]:
        if cell.value is not None:
            cell.fill      = fill
            cell.font      = font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = _border()


def _kpi_box(ws, row, col, label, value, bg="E8EAF6"):
    ws.cell(row,   col, label)
    ws.cell(row+1, col, value)
    ws.cell(row,   col).font      = Font(bold=True, size=10, color="555555")
    ws.cell(row+1, col).font      = Font(bold=True, size=14, color="1A237E")
    ws.cell(row+1, col).alignment = Alignment(horizontal="center")
    for r in [row, row+1]:
        ws.cell(r, col).fill   = PatternFill("solid", fgColor=bg)
        ws.cell(r, col).border = _border()
        ws.cell(r, col).alignment = Alignment(horizontal="center", vertical="center")


def _sheet_executive_summary(wb, kpis, ai, df):
    ws = wb.create_sheet("Executive Summary")

    # Title
    ws.merge_cells("A1:H1")
    ws["A1"] = "AUTOMATED SALES PERFORMANCE REPORT"
    ws["A1"].font      = Font(bold=True, size=18, color="1A237E")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A1"].fill      = PatternFill("solid", fgColor="E8EAF6")

    ws.merge_cells("A2:H2")
    ws["A2"] = f"Generated automatically on {datetime.now().strftime('%d %B %Y at %I:%M %p')}  |  Source: {INPUT_FILE}"
    ws["A2"].font      = Font(italic=True, size=10, color="888888")
    ws["A2"].alignment = Alignment(horizontal="center")

    # KPI boxes row
    ws.row_dimensions[4].height = 22
    ws.row_dimensions[5].height = 30
    _kpi_box(ws, 4, 1, "TOTAL REVENUE",    f"Rs {kpis['total_revenue']:,.0f}")
    _kpi_box(ws, 4, 2, "TOTAL UNITS SOLD", f"{kpis['total_units']:,}")
    _kpi_box(ws, 4, 3, "AVG ORDER VALUE",  f"Rs {kpis['avg_order_value']:,.0f}")
    _kpi_box(ws, 4, 4, "RETURN RATE",      f"{kpis['return_rate']:.1f}%",
             bg="FFF8E1" if kpis["return_rate"] > 5 else "E8F5E9")
    _kpi_box(ws, 4, 5, "TOP PRODUCT",
             kpis["by_product"].iloc[0]["product"], bg="F3E5F5")
    _kpi_box(ws, 4, 6, "TOP REGION",
             kpis["by_region"].iloc[0]["region"],   bg="E0F7FA")

    # AI Insights section
    ws.merge_cells("A7:H7")
    ws["A7"] = "AI-GENERATED INSIGHTS"
    ws["A7"].font      = Font(bold=True, size=13, color="1A237E")
    ws["A7"].alignment = Alignment(horizontal="left")

    insights = [
        ai.overall_health(kpis["total_revenue"], kpis["total_units"], kpis["avg_order_value"]),
        ai.growth_insight(kpis["last_month_rev"],   kpis["prev_month_rev"],   "Monthly Revenue"),
        ai.growth_insight(kpis["last_month_units"], kpis["prev_month_units"], "Units Sold"),
        ai.top_performer_insight(df, "product", "revenue"),
        ai.underperformer_insight(df, "product", "units_sold"),
        ai.trend_insight(kpis["monthly"].set_index("month")["revenue"]),
        ai.recommendation(df, "region", "revenue"),
    ]

    icons = ["📊", "📈", "📦", "🏆", "⚠️", "📉", "💡"]
    for i, (icon, insight) in enumerate(zip(icons, insights), 8):
        ws.merge_cells(f"A{i}:H{i}")
        ws[f"A{i}"] = f"  {icon}  {insight}"
        ws[f"A{i}"].font      = Font(size=11, color="222222")
        ws[f"A{i}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        bg = "F5F5F5" if i % 2 == 0 else "FFFFFF"
        ws[f"A{i}"].fill = PatternFill("solid", fgColor=bg)
        ws.row_dimensions[i].height = 30

    for col, w in enumerate([18,18,18,14,18,16,18,18], 1):
        ws.column_dimensions[get_column_letter(col)].width = w


def _sheet_monthly_trend(wb, kpis):
    ws   = wb.create_sheet("Monthly Trend")
    monthly = kpis["monthly"]

    ws["A1"] = "MONTHLY REVENUE & UNITS TREND"
    ws["A1"].font = Font(bold=True, size=14, color="1A237E")

    headers = ["Month", "Revenue (Rs)", "Units Sold", "Avg Revenue/Unit"]
    ws.append([""])
    ws.append(headers)
    _hdr(ws, 3, "1A237E")

    for _, row in monthly.iterrows():
        avg = row["revenue"] / row["units"] if row["units"] > 0 else 0
        ws.append([row["month"], round(row["revenue"],2), row["units"], round(avg,2)])
        for col in range(1, 5):
            ws.cell(ws.max_row, col).border    = _border()
            ws.cell(ws.max_row, col).alignment = Alignment(horizontal="center")

    # Line chart
    chart      = LineChart()
    chart.title = "Monthly Revenue Trend"
    chart.y_axis.title = "Revenue (Rs)"
    chart.x_axis.title = "Month"
    chart.style = 10
    chart.width = 20
    chart.height= 12

    data = Reference(ws, min_col=2, min_row=3, max_row=3+len(monthly))
    cats = Reference(ws, min_col=1, min_row=4, max_row=3+len(monthly))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "F3")

    for col, w in zip(["A","B","C","D"], [10, 16, 14, 18]):
        ws.column_dimensions[col].width = w


def _sheet_product_analysis(wb, kpis):
    ws = wb.create_sheet("Product Analysis")

    ws["A1"] = "PRODUCT PERFORMANCE ANALYSIS"
    ws["A1"].font = Font(bold=True, size=14, color="1A237E")

    ws.append([""])
    ws.append(["Product", "Revenue (Rs)", "Units Sold", "Returns", "Return Rate %", "Revenue Share %"])
    _hdr(ws, 3, "1A237E")

    total_rev = kpis["by_product"]["revenue"].sum()
    for _, row in kpis["by_product"].iterrows():
        ret_rate  = (row["returns"] / row["units"] * 100) if row["units"] > 0 else 0
        rev_share = (row["revenue"] / total_rev * 100)
        ws.append([
            row["product"],
            round(row["revenue"], 2),
            row["units"],
            row["returns"],
            round(ret_rate, 1),
            round(rev_share, 1)
        ])
        bg = "E8F5E9" if rev_share > 15 else ("FFF8E1" if rev_share > 8 else "FFEBEE")
        for col in range(1, 7):
            ws.cell(ws.max_row, col).fill      = PatternFill("solid", fgColor=bg)
            ws.cell(ws.max_row, col).border    = _border()
            ws.cell(ws.max_row, col).alignment = Alignment(horizontal="center")

    # Bar chart
    chart        = BarChart()
    chart.type   = "col"
    chart.title  = "Revenue by Product"
    chart.y_axis.title = "Revenue (Rs)"
    chart.style  = 10
    chart.width  = 20
    chart.height = 12

    n    = len(kpis["by_product"])
    data = Reference(ws, min_col=2, min_row=3, max_row=3+n)
    cats = Reference(ws, min_col=1, min_row=4, max_row=3+n)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "H3")

    for col, w in zip(["A","B","C","D","E","F"], [16,16,12,10,14,15]):
        ws.column_dimensions[col].width = w


def _sheet_raw_data(wb, df):
    ws = wb.create_sheet("Raw Data")
    ws.append(list(df.columns))
    _hdr(ws, 1, "37474F")
    for _, row in df.iterrows():
        ws.append(list(row))
        for col in range(1, len(df.columns)+1):
            ws.cell(ws.max_row, col).alignment = Alignment(horizontal="center")
            ws.cell(ws.max_row, col).border    = _border()
    ws.freeze_panes = "A2"
    for i, col in enumerate(df.columns, 1):
        ws.column_dimensions[get_column_letter(i)].width = 14


# ── MAIN ──────────────────────────────────────────────────────────────────────

def main():
    print("=" * 55)
    print("   AUTOMATED DATA REPORT GENERATOR")
    print("   By Yashwanth Eluri")
    print("=" * 55)

    df   = load_sales_data(INPUT_FILE)
    kpis = compute_kpis(df)
    ai   = AIInsightEngine()

    print("\nComputing KPIs...")
    print(f"  Total Revenue  : Rs {kpis['total_revenue']:,.0f}")
    print(f"  Total Units    : {kpis['total_units']:,}")
    print(f"  Return Rate    : {kpis['return_rate']:.1f}%")
    print(f"  Top Product    : {kpis['by_product'].iloc[0]['product']}")
    print(f"  Top Region     : {kpis['by_region'].iloc[0]['region']}")

    print("\nGenerating AI insights...")
    print("\nBuilding Excel report...")
    build_report(df, kpis, ai, OUTPUT_FILE)

    print(f"\nReport saved: {OUTPUT_FILE}")
    print("Sheets: Executive Summary | Monthly Trend | Product Analysis | Raw Data")
    print("\nDone in under 30 seconds!")
    print("=" * 55)


if __name__ == "__main__":
    main()
